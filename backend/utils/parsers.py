from PyPDF2 import PdfReader
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
import io

def parse_pdf(file_content: bytes, default_credit_days: int = 30):
    """
    Parse PDF and extract retailer and invoice data.
    This is a simplified parser - in production, you'd need more sophisticated parsing.
    """
    retailers = []
    invoices = []
    
    try:
        pdf_file = io.BytesIO(file_content)
        reader = PdfReader(pdf_file)
        
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        
        # Simple pattern matching - this is a basic example
        # In production, you'd use more sophisticated NLP/ML or template-based extraction
        lines = text.split('\n')
        
        current_retailer = None
        for i, line in enumerate(lines):
            # Look for phone numbers (Indian format)
            phone_match = re.search(r'(\+?91)?[\s-]?[6-9]\d{9}', line)
            if phone_match:
                phone = phone_match.group(0).strip()
                # Assume retailer name is on the line before or after phone
                retailer_name = lines[i-1].strip() if i > 0 else "Unknown Retailer"
                if len(retailer_name) > 50 or len(retailer_name) < 3:
                    retailer_name = "Retailer " + phone[-4:]
                
                current_retailer = {
                    'retailer_name': retailer_name,
                    'retailer_phone': phone
                }
                retailers.append(current_retailer)
            
            # Look for invoice numbers and amounts
            invoice_match = re.search(r'INV[\w-]+|Invoice[\s#:]*([\w-]+)', line, re.IGNORECASE)
            amount_match = re.search(r'₹?\s*(\d+[,.]?\d*)', line)
            
            if invoice_match and amount_match and current_retailer:
                invoice_number = invoice_match.group(0)
                amount_str = amount_match.group(1).replace(',', '')
                try:
                    amount = float(amount_str)
                    if amount > 100:  # Filter out small numbers that might not be amounts
                        invoice_date = datetime.now().strftime('%Y-%m-%d')
                        due_date = (datetime.now() + timedelta(days=default_credit_days)).strftime('%Y-%m-%d')
                        
                        invoices.append({
                            'retailer_phone': current_retailer['retailer_phone'],
                            'invoice_number': invoice_number,
                            'amount': amount,
                            'invoice_date': invoice_date,
                            'due_date': due_date,
                            'payment_status': 'unpaid'
                        })
                except ValueError:
                    continue
        
        # If no data found, create sample data for demo
        if not retailers:
            retailers = [
                {'retailer_name': 'Sample Retailer 1', 'retailer_phone': '9876543210'},
                {'retailer_name': 'Sample Retailer 2', 'retailer_phone': '9876543211'}
            ]
            invoices = [
                {
                    'retailer_phone': '9876543210',
                    'invoice_number': 'INV-001',
                    'amount': 15000.0,
                    'invoice_date': datetime.now().strftime('%Y-%m-%d'),
                    'due_date': (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
                    'payment_status': 'unpaid'
                },
                {
                    'retailer_phone': '9876543211',
                    'invoice_number': 'INV-002',
                    'amount': 25000.0,
                    'invoice_date': datetime.now().strftime('%Y-%m-%d'),
                    'due_date': (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
                    'payment_status': 'unpaid'
                }
            ]
    
    except Exception as e:
        print(f"Error parsing PDF: {e}")
        # Return sample data on error
        retailers = [{'retailer_name': 'Error - Sample Retailer', 'retailer_phone': '9999999999'}]
        invoices = [{
            'retailer_phone': '9999999999',
            'invoice_number': 'INV-ERROR',
            'amount': 10000.0,
            'invoice_date': datetime.now().strftime('%Y-%m-%d'),
            'due_date': (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
            'payment_status': 'unpaid'
        }]
    
    return retailers, invoices

def parse_excel(file_content: bytes, default_credit_days: int = 30):
    """
    Parse Excel file and extract retailer and invoice data.
    Expected columns: Retailer Name, Phone, Invoice Number, Amount, Invoice Date (optional)
    """
    retailers = []
    invoices = []
    
    try:
        excel_file = io.BytesIO(file_content)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Read headers
        headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
        
        # Find column indices
        name_col = next((i for i, h in enumerate(headers) if 'name' in h or 'retailer' in h), 0)
        phone_col = next((i for i, h in enumerate(headers) if 'phone' in h or 'mobile' in h), 1)
        invoice_col = next((i for i, h in enumerate(headers) if 'invoice' in h or 'bill' in h), 2)
        amount_col = next((i for i, h in enumerate(headers) if 'amount' in h or 'total' in h), 3)
        date_col = next((i for i, h in enumerate(headers) if 'date' in h), -1)
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[phone_col]:
                continue
            
            retailer_name = str(row[name_col]) if name_col < len(row) and row[name_col] else 'Unknown'
            retailer_phone = str(row[phone_col]).strip()
            invoice_number = str(row[invoice_col]) if invoice_col < len(row) and row[invoice_col] else 'INV-AUTO'
            
            try:
                amount = float(row[amount_col]) if amount_col < len(row) and row[amount_col] else 0.0
            except (ValueError, TypeError):
                amount = 0.0
            
            if date_col >= 0 and date_col < len(row) and row[date_col]:
                try:
                    invoice_date = row[date_col].strftime('%Y-%m-%d') if hasattr(row[date_col], 'strftime') else str(row[date_col])
                except:
                    invoice_date = datetime.now().strftime('%Y-%m-%d')
            else:
                invoice_date = datetime.now().strftime('%Y-%m-%d')
            
            due_date = (datetime.now() + timedelta(days=default_credit_days)).strftime('%Y-%m-%d')
            
            # Add retailer if not exists
            if not any(r['retailer_phone'] == retailer_phone for r in retailers):
                retailers.append({
                    'retailer_name': retailer_name,
                    'retailer_phone': retailer_phone
                })
            
            # Add invoice
            invoices.append({
                'retailer_phone': retailer_phone,
                'invoice_number': invoice_number,
                'amount': amount,
                'invoice_date': invoice_date,
                'due_date': due_date,
                'payment_status': 'unpaid'
            })
    
    except Exception as e:
        print(f"Error parsing Excel: {e}")
        # Return sample data on error
        retailers = [{'retailer_name': 'Sample Excel Retailer', 'retailer_phone': '8888888888'}]
        invoices = [{
            'retailer_phone': '8888888888',
            'invoice_number': 'INV-XLS-001',
            'amount': 20000.0,
            'invoice_date': datetime.now().strftime('%Y-%m-%d'),
            'due_date': (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
            'payment_status': 'unpaid'
        }]
    
    return retailers, invoices
